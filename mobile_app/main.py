import os
import shutil
from pathlib import Path
from functools import partial

import kivy
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.spinner import Spinner
from kivy.uix.popup import Popup
from kivy.uix.button import Button
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.scrollview import ScrollView

try:
    from openpyxl import load_workbook
except Exception:
    raise RuntimeError('openpyxl is required. Install with pip install openpyxl')


class JournalForm(GridLayout):
    def __init__(self, headers, workbook_path, **kwargs):
        super().__init__(cols=2, spacing=8, padding=8, size_hint_y=None, **kwargs)
        self.bind(minimum_height=self.setter('height'))
        self.headers = headers
        self.workbook_path = Path(workbook_path)
        self.inputs = {}

        for header in headers:
            lbl = Label(text=header, size_hint_y=None, height=40)
            # Choose widget type based on header name
            key = header.strip().upper()
            if key in ('CE/PE', 'CE/ PE', 'CE PE'):
                w = Spinner(values=['CE', 'PE'], size_hint_y=None, height=40)
            elif key in ('BUY/SELL', 'BUY / SELL'):
                w = Spinner(values=['BUY', 'SELL'], size_hint_y=None, height=40)
            elif 'QUANTITY' in key or key == 'QTY':
                w = TextInput(multiline=False, size_hint_y=None, height=40, input_filter='int')
            elif 'PRICE' in key or key == 'P&L' or 'P&L' in header:
                w = TextInput(multiline=False, size_hint_y=None, height=40, input_filter='float')
            elif key == 'DATE':
                # TextInput with a quick-fill Today button
                container = BoxLayout(size_hint_y=None, height=40)
                ti = TextInput(multiline=False)
                btn = Button(text='Today', size_hint_x=None, width=80)
                def set_today(_):
                    from datetime import date
                    ti.text = date.today().isoformat()
                btn.bind(on_release=set_today)
                container.add_widget(ti)
                container.add_widget(btn)
                w = container
            elif key in ('ENTRY TIME', 'EXIT TIME', 'TIME'):
                # TextInput with quick-fill Now button
                container = BoxLayout(size_hint_y=None, height=40)
                ti = TextInput(multiline=False)
                btn = Button(text='Now', size_hint_x=None, width=80)
                def set_now(_):
                    from datetime import datetime
                    ti.text = datetime.now().strftime('%H:%M')
                btn.bind(on_release=set_now)
                container.add_widget(ti)
                container.add_widget(btn)
                w = container
            else:
                w = TextInput(multiline=False, size_hint_y=None, height=40)
            self.inputs[header] = w
            self.add_widget(lbl)
            self.add_widget(w)

        btn_box = BoxLayout(size_hint_y=None, height=48)
        save_btn = Button(text='Save Entry')
        save_btn.bind(on_release=self.save_entry)
        export_btn = Button(text='Export to Downloads')
        export_btn.bind(on_release=self.export_workbook)
        btn_box.add_widget(save_btn)
        btn_box.add_widget(export_btn)
        self.add_widget(Label())
        self.add_widget(btn_box)

        # Bindings for auto-calculation of P&L
        def try_bind(key):
            if key in self.inputs and hasattr(self.inputs[key], 'bind'):
                self.inputs[key].bind(text=self._on_value_change)

        try_bind('ENTRY PRICE')
        try_bind('EXIT PRICE')
        try_bind('QUANTITY')
        try_bind('BUY/SELL')
        try_bind('P&L')

    def _on_value_change(self, instance, value):
        # Recalculate P&L if possible
        try:
            entry = float(self.inputs.get('ENTRY PRICE', TextInput()).text or 0)
            exitp = float(self.inputs.get('EXIT PRICE', TextInput()).text or 0)
            qty = float(self.inputs.get('QUANTITY', TextInput()).text or 0)
        except Exception:
            return
        pnl = (exitp - entry) * qty
        # If BUY/SELL indicates SELL, invert sign
        bs = self.inputs.get('BUY/SELL')
        if bs is not None and bs.text.strip().upper() == 'SELL':
            pnl = -pnl
        if 'P&L' in self.inputs:
            self.inputs['P&L'].text = str(round(pnl, 2))

    def save_entry(self, *args):
        values = []
        # validate and collect values
        errors = []
        for h in self.headers:
            w = self.inputs[h]
            val = ''
            if isinstance(w, Spinner):
                val = w.text or ''
            else:
                # container widgets (for DATE/TIME quick-fill) are BoxLayout
                if hasattr(w, 'children') and len(w.children) > 0 and isinstance(w.children[0], TextInput) is False:
                    # Try to find TextInput inside
                    ti = None
                    for c in w.children:
                        if isinstance(c, TextInput):
                            ti = c
                            break
                    val = ti.text if ti is not None else ''
                else:
                    val = w.text or ''
            values.append(val)
            # basic required-field checks for common columns
            key = h.strip().upper()
            if key in ('DATE','INSTRUMENT','BUY/SELL','ENTRY PRICE','QUANTITY') and (val is None or str(val).strip()==''):
                errors.append(f'{h} is required')
            if key in ('ENTRY PRICE','EXIT PRICE') and val:
                try:
                    float(str(val))
                except Exception:
                    errors.append(f'{h} must be a number')
            if key == 'QUANTITY' and val:
                try:
                    int(float(str(val)))
                except Exception:
                    errors.append(f'{h} must be an integer')
        if errors:
            self._show_popup('Validation error', '\n'.join(errors))
            return
        # Ensure workbook exists
        if not self.workbook_path.exists():
            App.get_running_app().show_message('Workbook not found: ' + str(self.workbook_path))
            return
        wb = load_workbook(filename=str(self.workbook_path), keep_vba=True)
        sheet_name = 'Sheet1' if 'Sheet1' in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        # Find next empty row after header (headers at row 9)
        row = ws.max_row + 1
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=row, column=col_idx, value=val)
        wb.save(filename=str(self.workbook_path))
        App.get_running_app().show_message('Saved to ' + str(self.workbook_path))
        self._show_popup('Saved', 'Entry saved successfully')
        # Clear inputs
        for inp in self.inputs.values():
            if isinstance(inp, Spinner):
                inp.text = ''
            else:
                inp.text = ''

    def _show_popup(self, title, message):
        box = BoxLayout(orientation='vertical', padding=8, spacing=8)
        box.add_widget(Label(text=message))
        btn = Button(text='OK', size_hint_y=None, height=40)
        box.add_widget(btn)
        popup = Popup(title=title, content=box, size_hint=(0.8, 0.4))
        btn.bind(on_release=popup.dismiss)
        popup.open()

    def export_workbook(self, *args):
        app = App.get_running_app()
        src = app.workbook_path
        if src is None or not src.exists():
            app.show_message('No workbook to export')
            return
        # Prefer SAF on Android (Android 11+ compatibility); fallback to Downloads copy
        if kivy_platform == 'android':
            # try SAF via ACTION_CREATE_DOCUMENT
            try:
                from jnius import autoclass, cast
                from android import activity

                Intent = autoclass('android.content.Intent')
                String = autoclass('java.lang.String')
                PythonActivity = autoclass('org.kivy.android.PythonActivity')
                activity_obj = PythonActivity.mActivity

                filename = f'JOURNAL_export_{int(Path().stat().st_mtime)}.xlsm'
                intent = Intent(Intent.ACTION_CREATE_DOCUMENT)
                intent.addCategory(Intent.CATEGORY_OPENABLE)
                intent.setType(String('application/vnd.ms-excel'))
                intent.putExtra(Intent.EXTRA_TITLE, filename)

                # store src path to write after on_activity_result
                self._saf_export_src = str(src)

                # bind result handler
                try:
                    activity.bind(on_activity_result=self._on_activity_result)
                except Exception:
                    pass

                # start activity
                activity_obj.startActivityForResult(intent, 1001)
                app.show_message('Choose location to save (SAF)')
                return
            except Exception:
                # fall back to Downloads copy below
                pass

        # Fallback: write to Downloads
        downloads = None
        try:
            if storagepath:
                downloads = Path(storagepath.get_downloads_dir())
        except Exception:
            downloads = None
        if not downloads:
            try:
                from jnius import autoclass
                Environment = autoclass('android.os.Environment')
                downloads = Path(Environment.getExternalStoragePublicDirectory(Environment.DIRECTORY_DOWNLOADS).getAbsolutePath())
            except Exception:
                downloads = Path.home() / 'Downloads'

        downloads.mkdir(parents=True, exist_ok=True)
        dest = downloads / (f'JOURNAL_export_{int(Path().stat().st_mtime)}.xlsm')
        try:
            shutil.copy2(str(src), str(dest))
            app.show_message('Exported to ' + str(dest))
        except Exception as e:
            app.show_message('Export failed: ' + str(e))

    def _on_activity_result(self, requestCode, resultCode, intent):
        # Handle SAF create document result
        try:
            from jnius import autoclass, cast
            from org import jnius
        except Exception:
            pass
        try:
            # RESULT_OK == -1
            if requestCode == 1001 and resultCode == -1 and intent is not None:
                uri = intent.getData()
                if uri is None:
                    App.get_running_app().show_message('No URI returned')
                    return

                # open file descriptor and write bytes
                PythonActivity = autoclass('org.kivy.android.PythonActivity')
                activity_obj = PythonActivity.mActivity
                pfd = activity_obj.getContentResolver().openFileDescriptor(uri, 'w')
                if pfd is None:
                    App.get_running_app().show_message('Could not open target URI')
                    return
                fd = pfd.getFileDescriptor()
                FileOutputStream = autoclass('java.io.FileOutputStream')
                fos = FileOutputStream(fd)
                # read source bytes
                src = getattr(self, '_saf_export_src', None)
                if not src:
                    App.get_running_app().show_message('No source file to export')
                    try:
                        fos.close()
                        pfd.close()
                    except Exception:
                        pass
                    return
                with open(src, 'rb') as f:
                    data = f.read()
                # write bytes
                # convert to Java byte[]
                jarray = jnius.cast('byte[]', data)
                fos.write(jarray)
                fos.close()
                pfd.close()
                App.get_running_app().show_message('Exported via SAF')
        except Exception as e:
            App.get_running_app().show_message('SAF export failed: ' + str(e))


class RootWidget(BoxLayout):
    def __init__(self, headers, workbook_path, **kwargs):
        super().__init__(orientation='vertical', spacing=8, padding=8, **kwargs)
        self.headers = headers
        self.workbook_path = workbook_path
        self.status = Label(size_hint_y=None, height=30)
        self.add_widget(self.status)

        scroll = ScrollView()
        form = JournalForm(headers, workbook_path, size_hint=(1, None))
        scroll.add_widget(form)
        self.add_widget(scroll)

    def set_status(self, text):
        self.status.text = text


class JournalApp(App):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        # Destination workbook path (writable location). On Android this is inside app user_data_dir.
        self.workbook_path = None
        self.packaged_workbook = Path(__file__).resolve().parent.parent / 'JOURNAL.xlsm'

    def build(self):
        # Ensure we have a writable copy in `user_data_dir`.
        data_dir = Path(self.user_data_dir)
        data_dir.mkdir(parents=True, exist_ok=True)
        dest = data_dir / 'JOURNAL.xlsm'
        self.workbook_path = dest
        # If dest missing, try to copy packaged workbook (bundled with app). If that fails, fall back to project root.
        if not dest.exists():
            copied = False
            try:
                if self.packaged_workbook.exists():
                    shutil.copy2(str(self.packaged_workbook), str(dest))
                    copied = True
                else:
                    # Try project root location
                    proj_root = Path(__file__).resolve().parents[1] / 'JOURNAL.xlsm'
                    if proj_root.exists():
                        shutil.copy2(str(proj_root), str(dest))
                        copied = True
            except Exception as e:
                print('Could not copy packaged workbook:', e)

        headers = self.read_headers()
        if not headers:
            root = BoxLayout(orientation='vertical')
            root.add_widget(Label(text='Could not read headers from JOURNAL.xlsm'))
            return root
        self.root_widget = RootWidget(headers, self.workbook_path)
        return self.root_widget

    def read_headers(self):
        if self.workbook_path is None or not self.workbook_path.exists():
            return None
        wb = load_workbook(filename=str(self.workbook_path), keep_vba=True)
        sheet_name = 'Sheet1' if 'Sheet1' in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        # headers on row 9 (1-indexed)
        header_row = list(ws.iter_rows(min_row=9, max_row=9, values_only=True))[0]
        headers = [h if h is not None else '' for h in header_row]
        # Filter out any empty trailing headers
        headers = [h for h in headers if str(h).strip() != '']
        return headers

    def show_message(self, text):
        self.root_widget.set_status(text)


if __name__ == '__main__':
    JournalApp().run()
